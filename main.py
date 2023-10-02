import tkinter
from tkinter import X, Button, Canvas, Label, Scrollbar, ttk, Entry, Checkbutton, messagebox
from tkcalendar import Calendar, DateEntry
import openpyxl
import os
from tkinter import filedialog
from PIL import Image, ImageTk
from fpdf import FPDF
import pandas as pd
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from collections import defaultdict

#from assetmgmt_functions import enter_data, add_asset_to_excel, add_item, clear_asset_entry_fields, update_item_list, get_print, escape_special_characters, generate_pdf, send_email, generate_consolidated_pdf


def enter_data():
    accepted = accepted_var.get()
    if accepted == "Accepted":
        # User Info
        name = name_label_e.get()
        empno = EmpNo_label_e.get()
        dept = Dept_label_e.get()
        email = Email_label_e.get()
        verifier = verifier_label_e.get()
        selected_date = cal.get()
        assets=[]
        items = Items_label_combobox.get()
        serialno= Serial_no_label_e.get()
        assetcode = asset_label_e.get() 
        remarks= Remarks_label_e.get()

        if items and serialno and assetcode and remarks:
            asset_list.append((items, serialno, assetcode, remarks))
            clear_asset_entry_fields()
            update_item_list()
        
        # Iterate through the asset list and add each item to the Excel sheet
        for asset in asset_list:
            items, serialno, assetcode, remarks = asset
            # Add the asset information to the Excel sheet
            add_asset_to_excel(name, dept, empno, selected_date, email, verifier, items, serialno,assetcode, remarks)

    else:
        messagebox.showinfo("Error", "Make sure to check the terms and conditions block :/")

def add_asset_to_excel(name, dept, empno, date, email, verifier, items, serialno,assetcode, remarks):
    filepath = r'C:\Users\Ojal\Documents\actual_asset_data.xlsx'
    if not os.path.exists(filepath):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        heading = ["Employee Name", "Department", "Emp No", "Verified Date", "Email", "Verifier Name","Items","Serial No","Asset Code","Remarks"]

        sheet.append(heading)
        workbook.save(filepath)

    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    sheet.append([name, dept, empno, date, email, verifier, items, serialno,assetcode, remarks])
    workbook.save(filepath)

def add_item():
    items = Items_label_combobox.get()
    serialno = Serial_no_label_e.get()
    assetcode = asset_label_e.get()
    remarks = Remarks_label_e.get()

    if items and serialno and assetcode and remarks:
        asset_list.append((items, serialno,assetcode, remarks))
        clear_asset_entry_fields()
        update_item_list()
    else:
        messagebox.showinfo("Error", "Please fill in all item details.")

def clear_asset_entry_fields():
    Items_label_combobox.set("")
    Serial_no_label_e.delete(0, tkinter.END)
    asset_label_e.delete(0,tkinter.END)
    Remarks_label_e.delete(0, tkinter.END)

def update_item_list():
    item_list.delete(0, tkinter.END)
    for asset in asset_list:
        items, serialno,assetcode, remarks = asset
        item_list.insert(tkinter.END, f"{items} - {serialno} - {assetcode} - {remarks}")

def get_print():
    name = name_label_e.get()
    empno = EmpNo_label_e.get()
    dept = Dept_label_e.get()
    title = title_label_combobox.get()
    email = Email_label_e.get()
    verifier = verifier_label_e.get()

    selected_date = cal.get()
    logo_path = r"C:\Users\Ojal\Documents\IT asset mngr\main\c1.jpg"

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    # logosize = logo.resize((100, 30))
    pdf.image(logo_path, x=10, y=10, w=30)
    pdf.cell(0, 10, "The Coca-Cola Bottling Company of Bahrain", 0, 1, "C")
    pdf.set_font("Arial", "", 12)

    pdf.line(10, pdf.get_y() + 5, 200, pdf.get_y() + 5)  # Add the horizontal line

    pdf.multi_cell(0, 10, txt=f"Name: {title} {name}     Employee Number: {empno}\n"
                              f"Department: {dept}     Email: {email}\n"
                              f"Verified Date: {selected_date}\n"
                              f"Verifier Name: {verifier}")

    pdf.cell(40, 10, "Item", border=1)
    pdf.cell(40, 10, "Serial Number", border=1)
    pdf.cell(40, 10, "Remarks", border=1)
    pdf.ln()  # Move to the next line

    for asset in asset_list:
        items, serialno, remarks = asset
        pdf.cell(40, 10, items, border=1)
        pdf.cell(40, 10, serialno, border=1)
        pdf.cell(40, 10, remarks, border=1)
        pdf.ln()

    filepath = filedialog.asksaveasfilename(defaultextension=".pdf")
    pdf.output(filepath)
    
    print(f"Failed to send email to {email}")

def escape_special_characters(s):
    if isinstance(s, set):
        s = str(s)  # Convert the set to a string
    return s.replace('\\', '\\\\').replace(')', '\\)').replace('(', '\\(').replace('\r', '\\r')

def generate_pdf(asset_data):
    pdf = FPDF()
    pdf.add_page()

    logo_path = r"C:\Users\Ojal\Documents\IT asset mngr\main\c1.jpg"

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    logosize = logo.resize((100, 30))
    pdf.image(logo_path, x=10, y=10, w=30)
    pdf.cell(0, 10, "The Coca-Cola Bottling Company of Bahrain", 0, 1, "C")
    pdf.set_font("Arial", "", 12)

    pdf.line(10, pdf.get_y() + 5, 200, pdf.get_y() + 5)  # Add the horizontal line

    pdf.multi_cell(0, 10, txt=f"\nName: {asset_data['Employee Name']}     Employee Number: {asset_data['Emp No']}\n"
                              f"Department: {asset_data['Department']}     Email: {asset_data['Email']}\n"
                              f"Verified Date: {asset_data['Verified Date']}\n"
                              f"Verifier Name: {asset_data['Verifier Name']}")

    pdf.cell(40, 10, "Item", border=1)
    pdf.cell(50, 10, "Asset Code", border=1)
    pdf.cell(50, 10, "Serial Number", border=1)
    pdf.cell(50, 10, "Remarks", border=1)
    pdf.ln()  # Move to the next line

    pdf.cell(40, 10, str(asset_data['Items']), border=1)
    pdf.cell(50, 10, str(asset_data['Asset Code']), border=1)
    pdf.cell(50, 10, str(asset_data['Serial No']), border=1)
    pdf.cell(50, 10, str(asset_data['Remarks']), border=1)
    pdf.ln()

    pdf_file_path = f"C:\\Users\\Ojal\\Documents\\IT asset mngr\\main\\new{asset_data['Emp No']}_asset.pdf"
    pdf.output(pdf_file_path)

    return pdf_file_path

def send_email():
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    smtp_username = "ojalcoke@gmail.com"
    smtp_password = "hyxa ytnw okbs wklr"

    excel_file = r'C:\Users\Ojal\Documents\actual_asset_data.xlsx'
    df = pd.read_excel(excel_file)

    # Create a dictionary to group assets by email
    assets_by_email = defaultdict(list)

    for index, row in df.iterrows():
        # Extract asset data from the Excel file
        asset_data = {
            'Employee Name': row['Employee Name'],
            'Emp No': row['Emp No'],
            'Department': row['Department'],
            'Verified Date': row['Verified Date'],
            'Email': row['Email'],
            'Verifier Name': row['Verifier Name'],
            'Items': row['Items'],
            'Serial No': row['Serial No'],
            'Asset Code': row['Asset Code'],
            'Remarks': row['Remarks']
        }

        # Group assets by email
        assets_by_email[asset_data['Email']].append(asset_data)

    for email, assets in assets_by_email.items():
        # Generate a consolidated PDF for each email group
        pdf_file_path = generate_consolidated_pdf(assets)

        # Create an email message
        msg = MIMEMultipart()
        msg['From'] = "ojalcoke@gmail.com"  # Replace with your Gmail email
        msg['To'] = email
        msg['Subject'] = "Your Assets PDF"


        with open(r'C:\Users\Ojal\Documents\IT asset mngr\main\new\template.html', "r") as template_file:
            email_content = template_file.read()


        email_content = email_content.replace("[Recipient Name]", asset_data['Employee Name'])
        email_content = email_content.replace("[PDF Link]", f"<a href='cid:{pdf_file_path}'>View PDF</a>")
       
    
        # Add PDF attachment
 
        with open(pdf_file_path, "rb") as f:
            attachment = MIMEApplication(f.read(), _subtype="pdf")
            attachment.add_header("Content-Disposition", "attachment", filename="Assets_Info.pdf")
            msg.attach(attachment)

        msg.attach(MIMEText(email_content, "html"))

        # Send the email
        try:
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(smtp_username, smtp_password)
            server.sendmail(smtp_username, email, msg.as_string())
            server.quit()
            print(f"Email sent to {email}")
        except Exception as e:
            print(f"Failed to send email to {email}: {str(e)}")

def generate_consolidated_pdf(assets):
    pdf = FPDF()
    pdf.add_page()

    logo_path = r"C:\Users\Ojal\Documents\IT asset mngr\main\c1.jpg"
    pdf.image(logo_path, x=10, y=10, w=30)
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "The Coca-Cola Bottling Company of Bahrain", 0, 1, "C")

    pdf.set_font("Arial", "", 12)

    pdf.cell(0, 10, "IT Asset Verification Form", 0, 1, "C")

    
    # Set font for the content
    pdf.set_font("Arial", "", 12)

    # Add a horizontal line below the header
    pdf.line(10, pdf.get_y() + 5, 200, pdf.get_y() + 5)
    pdf.ln(15)

    first_asset_data = assets[0]
    pdf.multi_cell(0, 10, txt=f"Name: {first_asset_data['Employee Name']}                Employee Number: {first_asset_data['Emp No']}\n"
                              f"Department: {first_asset_data['Department']}             Email: {first_asset_data['Email']}\n"
                              f"Verified Date: {first_asset_data['Verified Date']}\n"
                              f"Verifier Name: {first_asset_data['Verifier Name']}")

    # Add a table with asset details
    pdf.ln(10)
    pdf.cell(40, 10, "Items", border=1)
    pdf.cell(50, 10, "Asset Code", border=1)
    pdf.cell(50, 10, "Serial Number", border=1)
    pdf.cell(40, 10, "Remarks", border=1)
    pdf.ln()

    for asset_data in assets:
        # Assuming asset_data['Items'] is a string containing asset details
        pdf.cell(40, 10, asset_data['Items'], border=1)
        pdf.cell(50, 10, asset_data['Asset Code'], border=1)
        pdf.cell(50, 10, asset_data['Serial No'], border=1)
        pdf.cell(40, 10, asset_data['Remarks'], border=1)
        pdf.ln()

    pdf_file_path = r"C:\Users\Ojal\Documents\IT asset mngr\main\consolidated_assets.pdf"
    pdf.output(pdf_file_path)

    return pdf_file_path


def send_email_to_selected_person(selected_person, pdf_file_path):
    smtp_server = "smtp.gmail.com"  # Replace with your SMTP server
    smtp_port = 587  # Replace with your SMTP port
    smtp_username = "ojalcoke@gmail.com"  # Replace with your email
    smtp_password = "hyxa ytnw okbs wklr"  # Replace with your password

    # Create an email message
    msg = MIMEMultipart()
    msg['From'] = smtp_username
    msg['To'] = selected_person['Email']
    msg['Subject'] = "Your Assets PDF"

    # Add PDF attachment
    with open(pdf_file_path, "rb") as f:
        attachment = MIMEApplication(f.read(), _subtype="pdf")
        attachment.add_header("Content-Disposition", "attachment", filename="Assets_Info.pdf")
        msg.attach(attachment)

    # Send the email
    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_username, smtp_password)
        server.sendmail(smtp_username, selected_person['Email'], msg.as_string())
        server.quit()
        print(f"Email sent to {selected_person['Email']}")
    except Exception as e:
        print(f"Failed to send email to {selected_person['Email']}: {str(e)}")


def generate_and_send_pdf(selected_person,all_asset):
    # Generate the PDF for the selected person
    pdf_file_path = generate_consolidated_pdf_s(selected_person,all_asset)
    send_email_to_selected_person(selected_person, pdf_file_path)

def generate_consolidated_pdf_s(selected_person, assets):
    pdf = FPDF()
    pdf.add_page()

    logo_path = r"C:\Users\Ojal\Documents\IT asset mngr\main\c1.jpg"
    pdf.image(logo_path, x=10, y=10, w=30)
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "The Coca-Cola Bottling Company of Bahrain", 0, 1, "C")

    pdf.set_font("Arial", "", 12)

    pdf.cell(0, 10, "IT Asset Verification Form", 0, 1, "C")

    
    # Set font for the content
    pdf.set_font("Arial", "", 12)

    # Add a horizontal line below the header
    pdf.line(10, pdf.get_y() + 5, 200, pdf.get_y() + 5)
    pdf.ln(15)

    pdf.multi_cell(0, 10, txt=f"Name: {selected_person['Employee Name']}                Employee Number: {selected_person['Emp No']}\n"
                              f"Department: {selected_person['Department']}             Email: {selected_person['Email']}\n"
                              f"Verified Date: {selected_person['Verified Date']}\n"
                              f"Verifier Name: {selected_person['Verifier Name']}")

    # Add a table with asset details
    pdf.ln(10)
    pdf.cell(40, 10, "Items", border=1)
    pdf.cell(50, 10, "Asset Code", border=1)
    pdf.cell(50, 10, "Serial Number", border=1)
    pdf.cell(40, 10, "Remarks", border=1)
    pdf.ln()

    for asset_data in assets:
        # Assuming asset_data['Items'] is a string containing asset details
        pdf.cell(40, 10, asset_data['Items'], border=1)
        pdf.cell(50, 10, asset_data['Asset Code'], border=1)
        pdf.cell(50, 10, asset_data['Serial No'], border=1)
        pdf.cell(40, 10, asset_data['Remarks'], border=1)
        pdf.ln()

    pdf_file_path = r"C:\Users\Ojal\Documents\IT asset mngr\main\consolidated_assets.pdf"
    pdf.output(pdf_file_path)

    return pdf_file_path

def send_selected_pdf():
    selected_empno = EmpNo_label_e.get()
    print(selected_empno)

    if not selected_empno:  # Check for an empty string
        messagebox.showinfo("Error", "Please enter an Employee Number.")
        return

    excel_file = r'C:\Users\Ojal\Documents\actual_asset_data.xlsx'
    df = pd.read_excel(excel_file)

    selected_person = None
    all_assets = []

    for index, row in df.iterrows():
        excel_emp = str(row['Emp No'])
        if excel_emp == selected_empno:
            selected_person = {
                'Employee Name': row['Employee Name'],
                'Emp No': row['Emp No'],
                'Department': row['Department'],
                'Verified Date': row['Verified Date'],
                'Email': row['Email'],
                'Verifier Name': row['Verifier Name']
            }

            asset_data = {
                'Items': row['Items'],
                'Serial No': row['Serial No'],
                'Asset Code': row['Asset Code'],
                'Remarks': row['Remarks']
            }

            all_assets.append(asset_data)

    print(selected_person)

    if selected_person:
        generate_and_send_pdf(selected_person, all_assets)
    else:
        messagebox.showinfo("Error", "Employee Number not found")
        return

def read_email_template(template_name):
    template_path = f"email_templates/{template_name}.html"  # Assuming templates are in an 'email_templates' directory
    with open(template_path, "r") as template_file:
        email_content = template_file.read()
    return email_content


asset_list = []

window = tkinter.Tk()

# parent window for everything else, big window
window.title("Data Entry for IT Assets")

title = Label(window, text=' IT Assets Management System', bg=None, fg=None, font=('Inter', 20, 'bold'))
title.pack(fill=X)

frame = tkinter.Frame(window, padx=5, pady=20)  # creating a frame inside the window
frame.pack()  # layout managers(pack,grid)



logo = Image.open(r'C:\Users\Ojal\Documents\IT asset mngr\main\c1.jpg')
logosize = logo.resize((100, 30))
display = ImageTk.PhotoImage(logosize)
logo_label = ttk.Label(window, image=display)
logo_label.place(x=25, y=5)

# Label frames, sort of like groups that come under a frame(think in figma terms)
user_info_frame = tkinter.LabelFrame(frame, text="User Information")
user_info_frame.grid(row=0, column=0, padx=5, pady=5)

name_label = tkinter.Label(user_info_frame, text="Employee Name")
name_label.grid(row=0, column=0)

EmpNo_label = tkinter.Label(user_info_frame, text="Employee Number")
EmpNo_label.grid(row=0, column=1)

name_label_e = tkinter.Entry(user_info_frame)
EmpNo_label_e = tkinter.Entry(user_info_frame)
name_label_e.grid(row=1, column=0)
EmpNo_label_e.grid(row=1, column=1)

title_label = tkinter.Label(user_info_frame, text="Title")
title_label_combobox = ttk.Combobox(user_info_frame, values=["", "Mr", "Mrs."])
title_label.grid(row=0, column=2)
title_label_combobox.grid(row=1, column=2)

Dept_label = tkinter.Label(user_info_frame, text="Department")
Dept_label.grid(row=0, column=3)
Dept_label_e = tkinter.Entry(user_info_frame)
Dept_label_e.grid(row=1, column=3)

# Calendar for selecting the date
cal = DateEntry(user_info_frame, date_pattern="dd-mm-yyyy")
cal.grid(row=3, column=2)

Date_label = tkinter.Label(user_info_frame, text="Verified Date")
Date_label.grid(row=2, column=2)

Email_label = tkinter.Label(user_info_frame, text="Email Address")
Email_label.grid(row=2, column=0)
Email_label_e = tkinter.Entry(user_info_frame)
Email_label_e.grid(row=3, column=0)

verifier_label = tkinter.Label(user_info_frame, text="Verifier Name")
verifier_label.grid(row=2, column=1)
verifier_label_e = tkinter.Entry(user_info_frame)
verifier_label_e.grid(row=3, column=1)

for widget in user_info_frame.winfo_children():
    widget.grid_configure(padx=20, pady=5)

# Item frame
item_info_frame = tkinter.LabelFrame(frame, text="Item")
item_info_frame.grid(row=1, column=0, padx=20, pady=20, sticky="news")

Items_label = tkinter.Label(item_info_frame, text="Items")
Items_label.grid(row=0, column=0)
Items_label_combobox = ttk.Combobox(item_info_frame, values=["Laptop", "Desktop", "Scanner", "Printer", "Mobile Phone"])
Items_label_combobox.grid(row=1, column=0)

Serial_no_label = tkinter.Label(item_info_frame, text="Serial No")
Serial_no_label.grid(row=0, column=1)
Serial_no_label_e = tkinter.Entry(item_info_frame)
Serial_no_label_e.grid(row=1, column=1)

Remarks_label = tkinter.Label(item_info_frame, text="Remarks")
Remarks_label.grid(row=0, column=3)
Remarks_label_e = tkinter.Entry(item_info_frame)
Remarks_label_e.grid(row=1, column=3)

asset_label = tkinter.Label(item_info_frame, text="Asset Code")
asset_label.grid(row=0, column=2)
asset_label_e = tkinter.Entry(item_info_frame)
asset_label_e.grid(row=1, column=2)

# Create a listbox to display items
item_list = tkinter.Listbox(item_info_frame, selectmode=tkinter.SINGLE)
item_list.grid(row=2, column=0, columnspan=3, padx=5, pady=5)


add_item_button = tkinter.Button(item_info_frame, text="Add Item", command=add_item)
add_item_button.grid(row=2, column=3)

for widget in item_info_frame.winfo_children():
    widget.grid_configure(padx=23, pady=5, sticky="news")

terms_frame = tkinter.LabelFrame(frame, text="Terms & Conditions")
terms_frame.grid(row=2, column=0, padx=25, pady=10, sticky="news")

name = name_label_e.get()
accepted_var = tkinter.StringVar(value="Not Accepted")
# terms_check = tkinter.Checkbutton(terms_frame, text=f"I,{name} hereby acknowledge that above mentioned assets are "
#                                                   f"under my custody. I understand that this asset belongs to\n "
#                                                   f"The Coca-Cola Bottling Company of Bahrain(B.S.C) and is under my "
#                                                   f"possession for carrying out my office work.\n I hereby assure that  "
#                                                   f"I will take care of the assets of the compay to the best possible extend.",
#                                   variable=accepted_var, onvalue="Accepted", offvalue="Not Accepted")
# terms_check.grid(row=0, column=0, padx=5, pady=5)

button_print = tkinter.Button(window, text="Print", command=get_print)
button_print.place(relx=1.0, x=-30, y=5, anchor="ne")

button_frame = tkinter.Frame(frame)

button_frame.grid(row=3, column=0, columnspan=2)  # Columnspan to make it span two columns
button = tkinter.Button(button_frame, text="Enter data", command=enter_data, bg="white", fg="black")
button.grid(row=0, column=0, padx=10, pady=5)

send_email_button = tkinter.Button(button_frame, text="Send Email", command=send_email)
send_email_button.grid(row=0, column=1, padx=10, pady=5)

send_selected = tkinter.Button(button_frame,text="Send Selected",command = send_selected_pdf)
send_selected.grid(row=0,column = 2,padx =10, pady=5)

window.mainloop()
